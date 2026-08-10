#!/usr/bin/env python3
"""Apply the finalized TOEE Korean translation sheet to original loose game files.

The tool is deliberately non-destructive by default. It reads the finalized XLSX,
finds the matching original Co8/Temple+ files, replaces only translated visible
fields, restores ⟦CTRL:NN⟧ markers to their original C0 control characters, and
writes an overlay tree suitable for installation/testing.

Supported final-sheet file types: DLG, MES, TAB/TSV, JSON.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl이 필요합니다: python -m pip install openpyxl") from exc

CONTROL_MARKER_RE = re.compile(r"⟦CTRL:([0-9A-Fa-f]{2})⟧")
TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "cp1252", "latin-1")
JSON_PATH_TOKEN_RE = re.compile(r"\.([^\.\[]+)|\[(\d+)\]")


@dataclass(frozen=True)
class Translation:
    row_id: str
    source_set: str
    relative_path: str
    file_type: str
    entry_id: str
    text_slot: str
    english: str
    korean: str


@dataclass
class FileResult:
    source_set: str
    relative_path: str
    file_type: str
    targets: int = 0
    applied: int = 0
    skipped_mismatch: int = 0
    missing_entry: int = 0
    duplicate_entry: int = 0
    output_path: str = ""
    error: str = ""


def restore_controls(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        value = int(match.group(1), 16)
        if value <= 0x1F or value == 0x7F:
            return chr(value)
        return match.group(0)
    return CONTROL_MARKER_RE.sub(repl, text or "")


def escape_controls(text: str) -> str:
    out: list[str] = []
    for ch in text or "":
        code = ord(ch)
        if code <= 0x1F or code == 0x7F:
            out.append(f"⟦CTRL:{code:02X}⟧")
        else:
            out.append(ch)
    return "".join(out)


def read_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for enc in TEXT_ENCODINGS:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            pass
    return raw.decode("latin-1"), "latin-1"


def write_utf8_no_bom(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(text.encode("utf-8"))


def top_level_brace_spans(line: str) -> list[tuple[int, int, str]]:
    """Return (open_index, close_index, content) for top-level {...} groups.

    Nested braces are allowed inside a field (e.g. placeholders), matching the
    extractor's behavior. Escaped braces do not alter nesting.
    """
    spans: list[tuple[int, int, str]] = []
    i = 0
    n = len(line)
    while i < n:
        if line[i] != "{" or (i > 0 and line[i - 1] == "\\"):
            i += 1
            continue
        start = i
        i += 1
        depth = 1
        content_start = i
        while i < n and depth:
            ch = line[i]
            escaped = i > 0 and line[i - 1] == "\\"
            if ch == "{" and not escaped:
                depth += 1
            elif ch == "}" and not escaped:
                depth -= 1
                if depth == 0:
                    spans.append((start, i, line[content_start:i]))
                    i += 1
                    break
            i += 1
        else:
            return spans
    return spans


def replace_group(line: str, spans: list[tuple[int, int, str]], group_index: int, value: str) -> str:
    start, end, _ = spans[group_index]
    return line[: start + 1] + value + line[end:]


def newline_style(text: str) -> str:
    if "\r\n" in text:
        return "\r\n"
    if "\r" in text and "\n" not in text:
        return "\r"
    return "\n"


def split_keep_newlines(text: str) -> list[str]:
    """Split only on CRLF/CR/LF, never on C0 controls such as VT (0x0B)."""
    if not text:
        return []
    parts = re.split(r"(\r\n|\r|\n)", text)
    out: list[str] = []
    for i in range(0, len(parts), 2):
        body = parts[i]
        eol = parts[i + 1] if i + 1 < len(parts) else ""
        if body or eol:
            out.append(body + eol)
    return out


def line_body_and_eol(line: str) -> tuple[str, str]:
    if line.endswith("\r\n"):
        return line[:-2], "\r\n"
    if line.endswith("\n") or line.endswith("\r"):
        return line[:-1], line[-1]
    return line, ""


def _edge_ws(s: str) -> tuple[str, str, str]:
    m = re.fullmatch(r"([ \t]*)(.*?)([ \t]*)", s, flags=re.DOTALL)
    assert m is not None
    return m.group(1), m.group(2), m.group(3)


def source_match_and_replacement(original_visible: str, sheet_english: str, korean: str) -> tuple[bool, str]:
    """Validate source text and preserve source-only edge spaces/tabs.

    Google Sheets can trim edge spaces during intermediary cleanup. Those spaces
    are significant in composited MES fragments, so if the *only* difference is
    edge spaces/tabs, accept the source and carry its exact affixes into Korean.
    """
    marked = escape_controls(original_visible)
    expected = sheet_english or ""
    replacement = restore_controls(korean)
    if marked == expected:
        return True, replacement
    lp, core, rp = _edge_ws(marked)
    ep_l, ep_core, ep_r = _edge_ws(expected)
    if core == ep_core and not ep_l and not ep_r:
        return True, restore_controls(lp) + replacement.strip(" \t") + restore_controls(rp)
    return False, replacement


def safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def patch_dlg(text: str, trs: list[Translation], force: bool) -> tuple[str, Counter, list[dict[str, str]]]:
    wanted: dict[tuple[str, str], Translation] = {(t.entry_id, t.text_slot.upper()): t for t in trs}
    seen: Counter = Counter()
    stats = Counter()
    issues: list[dict[str, str]] = []
    out: list[str] = []

    for raw_line in split_keep_newlines(text):
        body, eol = line_body_and_eol(raw_line)
        spans = top_level_brace_spans(body)
        if len(spans) < 2:
            out.append(raw_line)
            continue
        entry_id = spans[0][2]
        line = body
        replacements: list[tuple[int, Translation]] = []
        for group_idx, slot in ((2, "B"), (1, "A")):
            tr = wanted.get((entry_id, slot))
            if tr and group_idx < len(spans):
                replacements.append((group_idx, tr))
        for group_idx, tr in replacements:
            seen[(tr.entry_id, tr.text_slot.upper())] += 1
            original = spans[group_idx][2]
            matched, replacement = source_match_and_replacement(original, tr.english, tr.korean)
            if not force and not matched:
                stats["mismatch"] += 1
                issues.append({
                    "row_id": tr.row_id,
                    "entry_id": tr.entry_id,
                    "slot": tr.text_slot,
                    "issue": "source_mismatch",
                    "source": escape_controls(original),
                    "sheet_english": tr.english,
                })
                continue
            line = replace_group(line, spans, group_idx, replacement if matched else restore_controls(tr.korean))
            spans = top_level_brace_spans(line)
            stats["applied"] += 1
        out.append(line + eol)

    for key, tr in wanted.items():
        count = seen[key]
        if count == 0:
            stats["missing"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "missing_entry"})
        elif count > 1:
            stats["duplicate"] += count - 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": f"duplicate_entry:{count}"})
    return "".join(out), stats, issues


def patch_mes(text: str, trs: list[Translation], force: bool) -> tuple[str, Counter, list[dict[str, str]]]:
    wanted: dict[str, Translation] = {t.entry_id: t for t in trs}
    seen: Counter = Counter()
    stats = Counter()
    issues: list[dict[str, str]] = []
    out: list[str] = []

    for raw_line in split_keep_newlines(text):
        body, eol = line_body_and_eol(raw_line)
        spans = top_level_brace_spans(body)
        if len(spans) < 2:
            out.append(raw_line)
            continue
        entry_id = spans[0][2]
        tr = wanted.get(entry_id)
        if not tr:
            out.append(raw_line)
            continue
        seen[entry_id] += 1
        original = spans[1][2]
        matched, replacement = source_match_and_replacement(original, tr.english, tr.korean)
        if not force and not matched:
            stats["mismatch"] += 1
            issues.append({
                "row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot,
                "issue": "source_mismatch", "source": escape_controls(original), "sheet_english": tr.english,
            })
            out.append(raw_line)
            continue
        line = replace_group(body, spans, 1, replacement if matched else restore_controls(tr.korean))
        stats["applied"] += 1
        out.append(line + eol)

    for entry_id, tr in wanted.items():
        count = seen[entry_id]
        if count == 0:
            stats["missing"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "missing_entry"})
        elif count > 1:
            stats["duplicate"] += count - 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": f"duplicate_entry:{count}"})
    return "".join(out), stats, issues


def parse_tab_rows(text: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(text, newline=""), delimiter="\t")
    return [row for row in reader]


def render_tab_rows(rows: list[list[str]], eol: str) -> str:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, delimiter="\t", lineterminator=eol, quoting=csv.QUOTE_MINIMAL)
    writer.writerows(rows)
    return buf.getvalue()


def patch_tab(text: str, trs: list[Translation], force: bool) -> tuple[str, Counter, list[dict[str, str]]]:
    rows = parse_tab_rows(text)
    eol = newline_style(text)
    by_key: dict[str, list[int]] = defaultdict(list)
    for idx, row in enumerate(rows):
        if row:
            by_key[row[0].strip()].append(idx)

    stats = Counter()
    issues: list[dict[str, str]] = []
    for tr in trs:
        m = re.fullmatch(r"COL(\d+)", tr.text_slot.upper())
        if not m:
            stats["missing"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "unsupported_tab_slot"})
            continue
        col = int(m.group(1)) - 1
        candidates = by_key.get(tr.entry_id, [])
        if not candidates:
            stats["missing"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "missing_entry"})
            continue
        if len(candidates) > 1:
            stats["duplicate"] += len(candidates) - 1
        applied_here = False
        for ridx in candidates:
            row = rows[ridx]
            if col >= len(row):
                continue
            original = row[col]
            matched, replacement = source_match_and_replacement(original, tr.english, tr.korean)
            if not force and not matched:
                continue
            row[col] = replacement if matched else restore_controls(tr.korean)
            stats["applied"] += 1
            applied_here = True
            break
        if not applied_here:
            stats["mismatch"] += 1
            source_preview = " | ".join(
                escape_controls(rows[r][col]) for r in candidates if col < len(rows[r])
            )
            issues.append({
                "row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot,
                "issue": "source_mismatch_or_column_missing", "source": source_preview, "sheet_english": tr.english,
            })
    return render_tab_rows(rows, eol), stats, issues


def json_path_tokens(path: str) -> list[str | int]:
    if not path.startswith("$"):
        raise ValueError(f"unsupported JSON path: {path}")
    tokens: list[str | int] = []
    pos = 1
    for m in JSON_PATH_TOKEN_RE.finditer(path, pos):
        if m.start() != pos:
            raise ValueError(f"unsupported JSON path segment: {path[pos:m.start()]}")
        tokens.append(m.group(1) if m.group(1) is not None else int(m.group(2)))
        pos = m.end()
    if pos != len(path):
        raise ValueError(f"unsupported JSON path tail: {path[pos:]}")
    return tokens


def get_json_path(obj: Any, path: str) -> Any:
    cur = obj
    for token in json_path_tokens(path):
        cur = cur[token]
    return cur


def set_json_path(obj: Any, path: str, value: str) -> None:
    tokens = json_path_tokens(path)
    if not tokens:
        raise ValueError("refusing to replace JSON document root")
    cur = obj
    for token in tokens[:-1]:
        cur = cur[token]
    cur[tokens[-1]] = value


def patch_json(text: str, trs: list[Translation], force: bool) -> tuple[str, Counter, list[dict[str, str]]]:
    obj = json.loads(text)
    stats = Counter()
    issues: list[dict[str, str]] = []
    for tr in trs:
        try:
            original = get_json_path(obj, tr.entry_id)
        except Exception as exc:
            stats["missing"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": f"missing_json_path:{exc}"})
            continue
        if not isinstance(original, str):
            stats["mismatch"] += 1
            issues.append({"row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "json_target_not_string"})
            continue
        matched, replacement = source_match_and_replacement(original, tr.english, tr.korean)
        if not force and not matched:
            stats["mismatch"] += 1
            issues.append({
                "row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot,
                "issue": "source_mismatch", "source": escape_controls(original), "sheet_english": tr.english,
            })
            continue
        set_json_path(obj, tr.entry_id, replacement if matched else restore_controls(tr.korean))
        stats["applied"] += 1
    trailing_newline = "\n" if text.endswith(("\n", "\r")) else ""
    return json.dumps(obj, ensure_ascii=False, indent=2) + trailing_newline, stats, issues


def load_translations(xlsx: Path, sheet_name: str | None = None) -> tuple[list[Translation], str]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    elif "TOEE_Translation_FILTERED_v2" in wb.sheetnames:
        ws = wb["TOEE_Translation_FILTERED_v2"]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    cols = {str(v): i for i, v in enumerate(header) if v is not None}
    required = ["RowID", "SourceSet", "RelativePath", "FileType", "EntryID", "TextSlot", "English", "Korean"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise ValueError(f"시트 필수 열 없음: {', '.join(missing)}")

    out: list[Translation] = []
    for row in rows:
        korean = row[cols["Korean"]]
        english = row[cols["English"]]
        if english in (None, ""):
            continue
        if korean in (None, ""):
            raise ValueError(f"번역 공란 발견: RowID={row[cols['RowID']]}")
        out.append(Translation(
            row_id=safe_text(row[cols["RowID"]]),
            source_set=safe_text(row[cols["SourceSet"]]),
            relative_path=safe_text(row[cols["RelativePath"]]).replace("\\", "/"),
            file_type=safe_text(row[cols["FileType"]]).upper(),
            entry_id=safe_text(row[cols["EntryID"]]),
            text_slot=safe_text(row[cols["TextSlot"]]),
            english=str(english),
            korean=str(korean),
        ))
    return out, ws.title


def latest_templeplus_tpdata() -> Path | None:
    local = os.environ.get("LOCALAPPDATA")
    if not local:
        return None
    base = Path(local) / "TemplePlus"
    apps = [p for p in base.glob("app-*") if p.is_dir() and (p / "tpdata").is_dir()]
    if not apps:
        return None

    def version_key(p: Path) -> tuple[int, ...]:
        parts: list[int] = []
        for x in p.name.removeprefix("app-").split("."):
            try:
                parts.append(int(x))
            except ValueError:
                parts.append(0)
        return tuple(parts)

    return sorted(apps, key=version_key, reverse=True)[0] / "tpdata"


def source_roots(game_root: Path, templeplus_root: Path | None) -> dict[str, Path]:
    roots = {
        "Co8Data": game_root / "data",
        "Co8Module": game_root / "modules" / "ToEE",
    }
    if templeplus_root:
        roots["TemplePlus"] = templeplus_root
    return roots


def staged_output_path(output_root: Path, source_set: str, rel: str) -> Path:
    relp = Path(rel.replace("/", os.sep))
    if source_set == "Co8Data":
        return output_root / "data" / relp
    if source_set == "Co8Module":
        return output_root / "modules" / "ToEE" / relp
    if source_set == "TemplePlus":
        return output_root / "TemplePlus" / "tpdata" / relp
    return output_root / source_set / relp


def in_place_output_path(root: Path, rel: str) -> Path:
    return root / Path(rel.replace("/", os.sep))


def make_backup(path: Path) -> None:
    bak = path.with_suffix(path.suffix + ".bak")
    if not bak.exists():
        shutil.copy2(path, bak)


def process_file(
    source_path: Path,
    output_path: Path,
    trs: list[Translation],
    file_type: str,
    force: bool,
    in_place: bool,
) -> tuple[FileResult, list[dict[str, str]]]:
    t0 = trs[0]
    result = FileResult(t0.source_set, t0.relative_path, file_type, targets=len(trs))
    if not source_path.is_file():
        result.error = f"source_file_missing: {source_path}"
        result.missing_entry = len(trs)
        return result, [{"row_id": t.row_id, "entry_id": t.entry_id, "slot": t.text_slot, "issue": "source_file_missing"} for t in trs]

    try:
        text, _encoding = read_text(source_path)
        if file_type == "DLG":
            patched, stats, issues = patch_dlg(text, trs, force)
        elif file_type == "MES":
            patched, stats, issues = patch_mes(text, trs, force)
        elif file_type in {"TAB", "TSV"}:
            patched, stats, issues = patch_tab(text, trs, force)
        elif file_type == "JSON":
            patched, stats, issues = patch_json(text, trs, force)
        else:
            result.error = f"unsupported_file_type:{file_type}"
            result.missing_entry = len(trs)
            return result, [{"row_id": t.row_id, "entry_id": t.entry_id, "slot": t.text_slot, "issue": result.error} for t in trs]

        result.applied = stats["applied"]
        result.skipped_mismatch = stats["mismatch"]
        result.missing_entry = stats["missing"]
        result.duplicate_entry = stats["duplicate"]

        if in_place:
            make_backup(source_path)
        write_utf8_no_bom(output_path, patched)
        result.output_path = str(output_path)
        return result, issues
    except Exception as exc:
        result.error = f"{type(exc).__name__}: {exc}"
        result.missing_entry = len(trs)
        return result, [{"row_id": t.row_id, "entry_id": t.entry_id, "slot": t.text_slot, "issue": result.error} for t in trs]


def validate_no_control_markers(paths: Iterable[Path]) -> list[str]:
    bad: list[str] = []
    for path in paths:
        try:
            raw = path.read_bytes()
        except OSError:
            continue
        if "⟦CTRL:".encode("utf-8") in raw:
            bad.append(str(path))
    return bad


def main() -> int:
    ap = argparse.ArgumentParser(description="TOEE 최종 한국어 번역을 원본 게임 텍스트 파일에 안전하게 재주입합니다.")
    ap.add_argument("--xlsx", type=Path, default=Path("TOEE_Translation_FILTERED_v2.xlsx"), help="최종 번역 XLSX")
    ap.add_argument("--sheet", default=None, help="시트명(기본: TOEE_Translation_FILTERED_v2 자동 선택)")
    ap.add_argument("--game-root", type=Path, default=Path.cwd(), help="TOEE 게임 루트(data, modules/ToEE가 있는 폴더)")
    ap.add_argument("--templeplus-root", type=Path, default=None, help="Temple+ tpdata 폴더. 생략 시 최신 설치 자동 감지")
    ap.add_argument("--output", type=Path, default=Path("TOEE_Korean_Patch_Output"), help="스테이징 출력 폴더")
    ap.add_argument("--in-place", action="store_true", help="원본 파일에 직접 적용(.bak 최초 1회 생성). 기본은 안전한 스테이징 출력")
    ap.add_argument("--force", action="store_true", help="영문 원본 대조가 달라도 강제 적용")
    args = ap.parse_args()

    xlsx = args.xlsx.resolve()
    game_root = args.game_root.resolve()
    output_root = args.output.resolve()
    tp_root = args.templeplus_root.resolve() if args.templeplus_root else latest_templeplus_tpdata()

    print("=== TOEE 한국어 최종 번역 재주입 ===")
    print("번역 시트:", xlsx)
    print("게임 루트:", game_root)
    print("Temple+:", tp_root if tp_root else "[감지 안 됨]")
    print("모드:", "원본 직접 적용 + .bak" if args.in_place else f"스테이징 생성 -> {output_root}")

    translations, sheet = load_translations(xlsx, args.sheet)
    print(f"시트: {sheet} / 번역 대상: {len(translations):,}행")

    grouped: dict[tuple[str, str, str], list[Translation]] = defaultdict(list)
    for tr in translations:
        grouped[(tr.source_set, tr.relative_path, tr.file_type)].append(tr)

    roots = source_roots(game_root, tp_root)
    file_results: list[FileResult] = []
    all_issues: list[dict[str, str]] = []
    written_paths: list[Path] = []

    for (source_set, rel, file_type), trs in sorted(grouped.items()):
        root = roots.get(source_set)
        if not root:
            res = FileResult(source_set, rel, file_type, targets=len(trs), missing_entry=len(trs), error="source_root_missing")
            file_results.append(res)
            for tr in trs:
                all_issues.append({"source_set": source_set, "relative_path": rel, "row_id": tr.row_id, "entry_id": tr.entry_id, "slot": tr.text_slot, "issue": "source_root_missing"})
            continue
        src = in_place_output_path(root, rel)
        dst = src if args.in_place else staged_output_path(output_root, source_set, rel)
        res, issues = process_file(src, dst, trs, file_type, args.force, args.in_place)
        file_results.append(res)
        if res.output_path:
            written_paths.append(Path(res.output_path))
        for issue in issues:
            issue.update({"source_set": source_set, "relative_path": rel, "file_type": file_type})
            all_issues.append(issue)

    summary = {
        "sheet": sheet,
        "translation_rows": len(translations),
        "files_targeted": len(grouped),
        "files_written": sum(1 for r in file_results if r.output_path),
        "rows_applied": sum(r.applied for r in file_results),
        "source_mismatches": sum(r.skipped_mismatch for r in file_results),
        "missing_entries_or_files": sum(r.missing_entry for r in file_results),
        "duplicate_entries": sum(r.duplicate_entry for r in file_results),
        "file_errors": sum(1 for r in file_results if r.error),
        "control_marker_files_remaining": [],
        "by_file_type": dict(Counter(t.file_type for t in translations)),
        "by_source_set": dict(Counter(t.source_set for t in translations)),
    }
    summary["control_marker_files_remaining"] = validate_no_control_markers(written_paths)

    report_dir = game_root if args.in_place else output_root
    report_dir.mkdir(parents=True, exist_ok=True)
    report = {
        "summary": summary,
        "files": [r.__dict__ for r in file_results],
        "issues": all_issues,
    }
    (report_dir / "PATCH_REPORT.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n=== 결과 ===")
    print(f"대상 파일: {summary['files_targeted']:,}")
    print(f"생성/수정 파일: {summary['files_written']:,}")
    print(f"적용 번역: {summary['rows_applied']:,} / {summary['translation_rows']:,}")
    print(f"원문 불일치로 보류: {summary['source_mismatches']:,}")
    print(f"누락 엔트리/파일: {summary['missing_entries_or_files']:,}")
    print(f"중복 엔트리: {summary['duplicate_entries']:,}")
    print(f"파일 오류: {summary['file_errors']:,}")
    print(f"CTRL 표식 잔존 파일: {len(summary['control_marker_files_remaining']):,}")
    print("보고서:", report_dir / "PATCH_REPORT.json")

    clean = (
        summary["rows_applied"] == summary["translation_rows"]
        and summary["source_mismatches"] == 0
        and summary["missing_entries_or_files"] == 0
        and summary["file_errors"] == 0
        and len(summary["control_marker_files_remaining"]) == 0
    )
    if clean:
        print("PATCH_BUILD_OK")
        return 0
    print("PATCH_BUILD_NEEDS_REVIEW")
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
